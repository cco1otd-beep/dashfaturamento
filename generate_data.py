#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Torre de Controle Logistica - OTD LOGISTICS
Gerador de data.js a partir das planilhas "lviagens" (base) e "lcrt" (Mercosul).

Uso:
    python3 generate_data.py lviagens_atualizado.xlsx [lcrt_atualizado.xlsx]

Se o segundo argumento (lcrt) nao for informado, o script procura
automaticamente por um arquivo "lcrt*.xlsx" na mesma pasta do lviagens
e, se nao encontrar, na pasta atual. Se nenhum lcrt for encontrado, o
script roda mesmo assim (sem o enriquecimento Mercosul).

Sempre que a OTD enviar uma nova planilha lviagens (e/ou lcrt), rode
este script dentro da pasta da Torre de Controle (mesma pasta do
index.html) para atualizar o arquivo data.js. Nao e necessario mexer
em nenhum outro arquivo (index.html / tv.html / style.css / chart.umd.min.js).

REGRAS DE NEGOCIO (definidas pela OTD):

1. Escopo: apenas o segmento AUTOPROPULSIONADO, identificado na
   planilha lviagens pela coluna "Grupo" = "RODANDO" (placas
   ficticias OTD-xxxx).

2. Cargas sem Nota Fiscal (coluna Q) sao descartadas da analise, pois
   nao representam carregamentos reais (romaneios administrativos/
   vazios) -- EXCETO quando o romaneio tem correspondencia na base
   lcrt (cargas do Mercosul, que sao faturadas via CRT e nao possuem
   Nota Fiscal brasileira).

3. Enriquecimento Mercosul: para romaneios que aparecem na base lcrt,
   quando o Frete Total (e/ou Frete Peso) do lviagens estiver zerado,
   usa-se o valor "Valor total CRT (R$)" da base lcrt como faturamento
   da carga. O cruzamento e feito pela coluna Romaneio (chave em
   ambas as planilhas). Quando uma linha do lcrt referencia mais de
   um romaneio na mesma celula (ex: "474921, 475006"), o valor do CRT
   e dividido igualmente entre os romaneios vinculados.
"""
import sys
import json
from datetime import datetime
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("Instale a dependencia antes de rodar: pip install openpyxl --break-system-packages")

SEGMENTO_GRUPO = "RODANDO"  # segmento autopropulsionado


def to_iso(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%dT%H:%M:%S")
    return None


def clean_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def clean_num(v):
    if v is None:
        return 0
    try:
        n = float(v)
        return 0 if n != n else n  # NaN guard
    except (TypeError, ValueError):
        return 0


def build_rota(carregamento, destino):
    c = clean_str(carregamento) or "?"
    d = clean_str(destino) or "?"
    return f"{c} → {d}"


def find_lcrt_file(lviagens_path):
    for folder in [lviagens_path.parent, Path(".")]:
        candidates = sorted(folder.glob("lcrt*.xlsx"))
        if candidates:
            return candidates[0]
    return None


def load_crt_map(lcrt_path):
    """Le a base lcrt e retorna {romaneio_int: valor_faturamento_rateado}."""
    wb = openpyxl.load_workbook(lcrt_path, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]

    def col(name):
        return header.index(name)

    rom_idx = col("Romaneio")
    val_idx = col("Valor total CRT (R$)")

    crt_map = defaultdict(float)
    linhas_lidas = 0
    linhas_sem_romaneio = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        rom_raw = row[rom_idx]
        if not rom_raw:
            linhas_sem_romaneio += 1
            continue
        valor = row[val_idx] or 0
        # algumas linhas do CRT referenciam mais de um romaneio na mesma celula
        partes = [p.strip() for p in str(rom_raw).split(",") if p.strip()]
        ids_validos = []
        for p in partes:
            try:
                ids_validos.append(int(float(p)))
            except (TypeError, ValueError):
                pass
        if not ids_validos:
            continue
        rateio = valor / len(ids_validos)
        for rid in ids_validos:
            crt_map[rid] += rateio
        linhas_lidas += 1

    return crt_map, linhas_lidas, linhas_sem_romaneio


def main():
    if len(sys.argv) < 2:
        candidates = [p for p in Path(".").glob("*.xlsx") if not p.name.lower().startswith("lcrt")]
        if not candidates:
            sys.exit("Informe o caminho do arquivo xlsx: python3 generate_data.py lviagens.xlsx [lcrt.xlsx]")
        src = candidates[0]
    else:
        src = Path(sys.argv[1])

    if not src.exists():
        sys.exit(f"Arquivo nao encontrado: {src}")

    lcrt_src = Path(sys.argv[2]) if len(sys.argv) > 2 else find_lcrt_file(src)
    crt_map = {}
    crt_linhas = crt_sem_romaneio = 0
    if lcrt_src and lcrt_src.exists():
        crt_map, crt_linhas, crt_sem_romaneio = load_crt_map(lcrt_src)
        print(f"lcrt encontrado: {lcrt_src.name} ({crt_linhas} linhas com romaneio, {crt_sem_romaneio} sem romaneio, {len(crt_map)} romaneios distintos)")
    else:
        print("AVISO: nenhum arquivo lcrt encontrado -- cargas do Mercosul com faturamento zerado NAO serao enriquecidas.")

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]

    def col(name):
        return header.index(name)

    idx = {
        "grupo": col("Grupo"),
        "romaneio": col("Nº Romaneio"),
        "conferencia": col("Conferência Romaneio"),
        "documento": col("Documento"),
        "cliente": col("Cliente"),
        "vazioDe": col("Vazio de"),
        "carregamento": col("Carregamento"),
        "destino": col("Destino"),
        "freteTotal": col("Frete Total"),
        "kmVazio": col("KM Vazio"),
        "kmCarregado": col("KM Carregado"),
        "placa": col("Placa Tração"),
        "motorista": col("Motorista"),
        "fretePeso": col("Frete Peso"),
        "notaFiscal": col("Nota Fiscal"),
        "dtSolicitacao": col("Dt. Carga Solicitação (I)"),
        "dtPrev": col("Dt. Prev. (C)"),
        "dtCargaI": col("Dt. Carga (I)"),
        "dtCargaT": col("Dt. Carga (T)"),
        "dtDescargaI": col("Dt. Descarga (I)"),
        "dtDescargaResT": col("Dt. Descarga Res. (T)"),
    }

    records = []
    skipped_other_segment = 0
    skipped_sem_nf_sem_crt = 0
    enriquecidos_crt = 0
    valor_enriquecido_total = 0.0

    for row in ws.iter_rows(min_row=2, values_only=True):
        grupo = clean_str(row[idx["grupo"]])
        if grupo != SEGMENTO_GRUPO:
            skipped_other_segment += 1
            continue

        romaneio = row[idx["romaneio"]]
        nota_fiscal = clean_str(row[idx["notaFiscal"]])
        crt_valor = crt_map.get(romaneio)

        # regra: sem Nota Fiscal E sem correspondencia no CRT -> nao e carga real, descarta
        if not nota_fiscal and crt_valor is None:
            skipped_sem_nf_sem_crt += 1
            continue

        dt_carga_i = row[idx["dtCargaI"]]
        dt_descarga_i = row[idx["dtDescargaI"]]
        dt_solicitacao = row[idx["dtSolicitacao"]]

        if not isinstance(dt_carga_i, datetime):
            status = "nao_iniciado"
        elif not isinstance(dt_descarga_i, datetime):
            status = "em_transito"
        else:
            status = "concluido"

        # mes de referencia para filtros: usa Dt.Carga(I); se nao houver,
        # usa a data de solicitacao (viagem ainda nao iniciada)
        ref_date = dt_carga_i if isinstance(dt_carga_i, datetime) else dt_solicitacao
        mes_ref = ref_date.strftime("%Y-%m") if isinstance(ref_date, datetime) else None

        placa = clean_str(row[idx["placa"]])

        frete = round(clean_num(row[idx["freteTotal"]]), 2)
        frete_peso = round(clean_num(row[idx["fretePeso"]]), 2)

        veio_do_crt = False
        if crt_valor is not None and frete == 0:
            frete = round(crt_valor, 2)
            veio_do_crt = True
        if crt_valor is not None and frete_peso == 0:
            frete_peso = round(crt_valor, 2)
            veio_do_crt = True
        if veio_do_crt:
            enriquecidos_crt += 1
            valor_enriquecido_total += crt_valor

        rec = {
            "id": romaneio,
            "conf": clean_str(row[idx["conferencia"]]),
            "doc": clean_str(row[idx["documento"]]),
            "cliente": clean_str(row[idx["cliente"]]),
            "origem": clean_str(row[idx["vazioDe"]]),
            "carreg": clean_str(row[idx["carregamento"]]),
            "destino": clean_str(row[idx["destino"]]),
            "rota": build_rota(row[idx["carregamento"]], row[idx["destino"]]),
            "frete": frete,
            "fretePeso": frete_peso,
            "kmVazio": round(clean_num(row[idx["kmVazio"]]), 1),
            "kmCarreg": round(clean_num(row[idx["kmCarregado"]]), 1),
            "placa": placa,
            "otd": bool(placa and placa.upper().startswith("OTD")),
            "motorista": clean_str(row[idx["motorista"]]),
            "nf": nota_fiscal,
            "mercosul": crt_valor is not None,
            "crt": veio_do_crt,
            "dtSol": to_iso(dt_solicitacao),
            "dtPrev": to_iso(row[idx["dtPrev"]]),
            "dtCargaI": to_iso(dt_carga_i),
            "dtCargaT": to_iso(row[idx["dtCargaT"]]),
            "dtDescI": to_iso(dt_descarga_i),
            "dtDescT": to_iso(row[idx["dtDescargaResT"]]),
            "status": status,
            "mesRef": mes_ref,
        }
        records.append(rec)

    meta = {
        "geradoEm": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "arquivoOrigem": src.name,
        "arquivoCrt": lcrt_src.name if (lcrt_src and lcrt_src.exists()) else None,
        "totalRegistros": len(records),
        "segmento": SEGMENTO_GRUPO,
        "registrosIgnoradosOutroSegmento": skipped_other_segment,
        "registrosIgnoradosSemNFSemCRT": skipped_sem_nf_sem_crt,
        "registrosEnriquecidosCRT": enriquecidos_crt,
        "valorEnriquecidoCRT": round(valor_enriquecido_total, 2),
    }

    out_path = Path(__file__).parent / "data.js"
    payload = "window.OTD_DATA = " + json.dumps(records, ensure_ascii=False) + ";\n"
    payload += "window.OTD_META = " + json.dumps(meta, ensure_ascii=False) + ";\n"
    out_path.write_text(payload, encoding="utf-8")

    print(f"OK: {len(records)} viagens (segmento {SEGMENTO_GRUPO}) gravadas em {out_path}")
    print(f"Ignorados (outros segmentos): {skipped_other_segment}")
    print(f"Ignorados (sem Nota Fiscal e sem CRT): {skipped_sem_nf_sem_crt}")
    print(f"Enriquecidos via CRT (Mercosul, faturamento antes zerado): {enriquecidos_crt} · R$ {valor_enriquecido_total:,.2f}")


if __name__ == "__main__":
    main()
